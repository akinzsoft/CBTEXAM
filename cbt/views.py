from django.shortcuts import redirect, render,get_object_or_404
from rest_framework.decorators import api_view
from django.contrib.auth.decorators import login_required
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from django.http import JsonResponse
#from django.contrib.auth.models import User
#from django.contrib.auth import authenticate, login
from django.db import transaction
from django.contrib.auth.hashers import make_password
from django.db.models import Max
from rest_framework import status
from rest_framework.decorators import api_view, permission_classes
from rest_framework import viewsets
from .models import Profile,User
from .serializer import questionserial,LoginSerializer,ScholarshipApplicantSerializer,RegisterSerializer,courseserial
from rest_framework.permissions import AllowAny
from django.utils.decorators import method_decorator
from django.views.decorators.csrf import csrf_exempt
import logging
from django.shortcuts import render, redirect
from .forms import AvatarForm,ProfileUpdateForm,ScholarshipApplicantForm
from django.contrib import messages
from django.utils import timezone
from django.contrib.auth import authenticate, login,logout,update_session_auth_hash
from django.contrib.auth.forms import PasswordChangeForm
from django.urls import reverse
import random
from django.db.models import Sum
import openpyxl
from django.http import HttpResponse 
from io import BytesIO

# Create your views here.
def home(request):
    if request.user.is_authenticated:   # ✅ check if logged in
        fname = request.user.first_name
        return render(request, 'index.html', {"fname": fname})
    else:
        return redirect('login/')

    

def login_page(request):
    return render(request, 'pages-sign-in.html')


def reg(request):
    if request.user.is_authenticated:  # check if logged in
        try:
            profile = request.user.profile  # get the related Profile
        except Profile.DoesNotExist:
            return redirect('no-access/')  # no profile, deny access

        if profile.role == "Admin":  # only Admin can access
            fname = request.user.first_name
            return render(request, 'pages-sign-up.html', {"fname": fname})
        else:
            return redirect('home')  # logged-in but not admin
    else:
        return redirect('login/')  # not logged in

# def reg(request):
#     if request.user.is_authenticated:   # ✅ check if logged in
#         fname = request.user.first_name
#         return render(request, 'pages-sign-up.html', {"fname": fname})
#     else:
#         return redirect('login/')
     




def login_user(request):
    error_message = None

    if request.method == "POST":
        username = request.POST.get('username')
        password = request.POST.get('password')

        if not username or not password:
            error_message = "Username and password are required."
        else:
            # Authenticate user
            user = authenticate(request, username=username, password=password)
            if user is not None:
                login(request, user)  # ✅ works with Django HttpRequest
                return redirect('home')
            else:
                error_message = "Invalid username or password."

    # Render login form with error message (if any)
    return render(request, 'pages-sign-in.html', {'error_message': error_message})


def register(request):
    if request.user.is_authenticated:
        if request.method == 'POST':
            serializer = RegisterSerializer(data=request.POST)
            if serializer.is_valid():
                serializer.save()
                messages.success(request, 'User registered successfully!')
                return redirect('register')
            else:
                for field, errs in serializer.errors.items():
                    for err in errs:
                        messages.error(request, f"{field}: {err}")
                return redirect('register')
        else:  # GET request
            return render(request, 'pages-sign-up.html')
    else:
        return redirect('home')
  
from django.shortcuts import render, redirect
from .forms import AvatarForm, ProfileUpdateForm
from .models import Profile,Course, Question,ScholarshipProgram, CBTExam,ScholarshipApplicant,StudentCBTAnswer


def profile(request):
    if not request.user.is_authenticated:
        return redirect('home')

    profile_obj, created = Profile.objects.get_or_create(user=request.user)

    if request.method == 'POST':
        avatar_form = AvatarForm(request.POST, request.FILES, instance=profile_obj)
        profile_form = ProfileUpdateForm(request.POST, instance=profile_obj)

        if avatar_form.is_valid() and profile_form.is_valid():
            # Save avatar
            avatar_form.save()

            # Update User fields
            request.user.first_name = profile_form.cleaned_data['first_name']
            request.user.last_name = profile_form.cleaned_data['last_name']
            request.user.email = profile_form.cleaned_data['email']
            request.user.save()

            profile_form.save()
            return redirect('profile')
    else:
        avatar_form = AvatarForm(instance=profile_obj)
        profile_form = ProfileUpdateForm(instance=profile_obj, initial={
            'first_name': request.user.first_name,
            'last_name': request.user.last_name,
            'email': request.user.email,
        })

    return render(request, "pages-profile.html", {
        'profile': profile_obj,
        'fname': request.user.first_name,
        'avatar_form': avatar_form,
        'profile_form': profile_form,
    })
    
    
@login_required
def users_list(request):
    # Only allow Admins
    if hasattr(request.user, 'profile') and request.user.profile.role == "Admin":
        users = User.objects.filter(is_staff=True).order_by('date_joined')  # Only staff users
        context = {
            'users': users,
            'now': timezone.now()
        }
        return render(request, 'staff.html', context)
    else:
        return redirect('home')  # Redirect non-admins to a "no access" page

@login_required
def update_role(request, user_id):
    # Ensure logged-in user has a profile and is allowed
    if not hasattr(request.user, "profile") or request.user.profile.role not in ["Admin", "Moderator", "User"]:
        messages.error(request, "You are not authorized to change roles.")
        return redirect("users_list")

    user = get_object_or_404(User, id=user_id)

    # Prevent changing your own role
    if user == request.user:
        messages.error(request, "You cannot change your own role.")
        return redirect("users_list")

    if request.method == "POST":
        role = request.POST.get("role")
        
        # Check if user has profile
        if hasattr(user, "profile"):
            # Update existing profile
            user.profile.role = role
            user.profile.save()
        else:
            # Create new profile with only user and role
            Profile.objects.create(user=user, role=role)

        messages.success(request, f"Role for {user.username} updated to {role}.")

    return redirect("users_list")


@login_required
def toggle_account_status(request, user_id):
    if request.user.profile.role not in ["Admin", "Moderator"]:
        messages.error(request, "You are not authorized to change account status.")
        return redirect("users_list")

    user = get_object_or_404(User, id=user_id)

    if user == request.user:
        messages.error(request, "You cannot disable your own account.")
        return redirect("users_list")

    # Toggle active status
    user.is_active = not user.is_active
    user.save()
    status = "enabled" if user.is_active else "disabled"
    messages.success(request, f"Account for {user.username} has been {status}.")
    return redirect("users_list")


@login_required
def remove_user(request, user_id):
    if request.user.profile.role != "Admin":  # Only Admins can delete
        messages.error(request, "Only admins can remove accounts.")
        return redirect("users_list")

    user = get_object_or_404(User, id=user_id)

    if user == request.user:
        messages.error(request, "You cannot remove your own account.")
        return redirect("users_list")

    # DELETE user permanently
    username = user.username
    user.delete()
    messages.success(request, f"User {username} has been permanently removed.")
    return redirect("users_list")


@login_required
def cbt_setup(request):
    """
    Display all courses and the total number of questions in each.
    """
    courses = Course.objects.all()
    return render(request, "cbt_setup.html", {"courses": courses})


# ========================
# Add Course View
# ========================
@login_required
def add_course(request):
    """
    Add a new course. Only Admin and Moderator can access.
    """
    # Check role
    if not hasattr(request.user, 'profile') or request.user.profile.role not in ["Admin", "Moderator"]:
        messages.error(request, "You do not have permission to access this page.")
        return redirect("no-access")  # or any page you prefer

    if request.method == "POST":
        title = request.POST.get("title")
        if title:
            Course.objects.create(title=title)
            messages.success(request, f"Course '{title}' added successfully!")
            return redirect("cbt_setup")
        else:
            messages.error(request, "Course title cannot be empty.")
    
    return render(request, "add_course.html")


# ========================
# Add Question View
# ========================
# Reusable decorator
from functools import wraps

def admin_moderator_required(view_func):
    @wraps(view_func)
    def wrapper(request, *args, **kwargs):
        if request.user.is_authenticated:
            if hasattr(request.user, 'profile') and request.user.profile.role in ["Admin", "Moderator"]:
                return view_func(request, *args, **kwargs)
        messages.error(request, "You do not have permission to access this page.")
        return redirect("home")  # redirect non-authorized users
    return wrapper

# Apply decorator to the view
@login_required
@admin_moderator_required
def add_question(request):
    """
    Add a new question to a selected course. Only Admin and Moderator can access.
    """
    courses = Course.objects.all()

    if request.method == "POST":
        course_id = request.POST.get("course")
        text = request.POST.get("text")
        option_a = request.POST.get("option_a")
        option_b = request.POST.get("option_b")
        option_c = request.POST.get("option_c")
        option_d = request.POST.get("option_d")
        correct_option = request.POST.get("correct_option")

        course = get_object_or_404(Course, id=course_id)
        Question.objects.create(
            course=course,
            text=text,
            option_a=option_a,
            option_b=option_b,
            option_c=option_c,
            option_d=option_d,
            correct_option=correct_option,
        )
        messages.success(request, "Question added successfully!")
        return redirect("cbt_setup")

    return render(request, "add_question.html", {"courses": courses})
@login_required
def view_questions(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    questions = course.questions.all()
    return render(request, "view_questions.html", {"course": course, "questions": questions})


def edit_question(request, question_id):
    # Fetch the question
    question = get_object_or_404(Question, id=question_id)
    
    if request.method == 'POST':
        # Update question fields from the form
        question.text = request.POST.get('text', '').strip()
        question.option_a = request.POST.get('option_a', '').strip()
        question.option_b = request.POST.get('option_b', '').strip()
        question.option_c = request.POST.get('option_c', '').strip()
        question.option_d = request.POST.get('option_d', '').strip()
        question.correct_option = request.POST.get('correct_option', '').strip()
        question.mrk = request.POST.get('mrk', '').strip()
        
        # Save the changes
        question.save()
        
        # Redirect back to the question list page
        return redirect(reverse('view_questions', args=[question.course.id]))
    
    # GET request – render the edit form
    return render(request, 'edit_question.html', {'question': question})

@login_required
def delete_question(request, question_id):
    question = get_object_or_404(Question, id=question_id)
    course_id = question.course.id
    if request.method == "POST":
        question.delete()
        messages.success(request, "Question deleted successfully!")
    return redirect("view_questions", course_id=course_id)



@login_required
def dd_question(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    if request.method == "POST":
        text = request.POST.get("text")
        option_a = request.POST.get("option_a")
        option_b = request.POST.get("option_b")
        option_c = request.POST.get("option_c")
        option_d = request.POST.get("option_d")
        correct_option = request.POST.get("correct_option")
        ansmrk = request.POST.get("ansmrk")

        Question.objects.create(
            course=course,
            text=text,
            option_a=option_a,
            option_b=option_b,
            option_c=option_c,
            option_d=option_d,
            correct_option=correct_option,
            mrk=ansmrk
        )

        messages.success(request, "Question added successfully!")
        return redirect("view_questions", course_id=course.id)

    return render(request, "addques.html", {"course": course})


def setup_scholarship(request):
    cbt_exams = CBTExam.objects.all()
    courses = Course.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title').strip()
        passing_score = request.POST.get('passing_score')
        exam_date = request.POST.get('exam_date')
        cbt_exam_id = request.POST.get('cbt_exam')
        course_id = request.POST.get('course')
        requirements = request.POST.get('requirements').strip()

        cbt_exam = CBTExam.objects.get(id=cbt_exam_id)

        if course_id == 'all':
            course = None
        else:
            course = Course.objects.get(id=course_id)

        ScholarshipProgram.objects.create(
            title=title,
            passing_score=passing_score,
            exam_date=exam_date,
            cbt_exam=cbt_exam,
            course=course,
            requirements=requirements
        )

        return redirect('scholarship_list')  # replace with your list view URL

    context = {
        'cbt_exams': cbt_exams,
        'courses': courses,
        'now': timezone.now(),
    }
    return render(request, 'sdash.html', context)


def add_scholarship(request):
    """
    View to add a new scholarship program.
    """
    cbt_exams = CBTExam.objects.all()
    courses = Course.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        passing_score = request.POST.get('passing_score')
        exam_date = request.POST.get('exam_date')
        cbt_exam_id = request.POST.get('cbt_exam')
        #course_id = request.POST.get('course')
        requirements = request.POST.get('requirements', '').strip()
        tquestion=request.POST.get('tquestion')
        exam_duration=request.POST.get('exam_duration')
        exam_end=request.POST.get('exam_end')
        exam_start=request.POST.get('exam_start')
        # Validate CBT exam
        cbt_exam = get_object_or_404(CBTExam, id=cbt_exam_id)

        # Handle course selection
        #if course_id == 'all':
        #    course = None  # Applies to all courses
        #else:
        #    course = get_object_or_404(Course, id=course_id)

        # Create the scholarship
        ScholarshipProgram.objects.create(
            title=title,
            passing_score=passing_score,
            exam_date=exam_date,
            cbt_exam=cbt_exam,
            #course=course,
            requirements=requirements,
            examduration=exam_duration,
            endtime=exam_end,
            starttime=exam_start
        )

        # Redirect to the scholarship list page
        return redirect('scholarship_list')

    context = {
        'cbt_exams': cbt_exams,
        #'courses': courses,
        'now': timezone.now(),
    }
    return render(request, 'setup_scholarship.html', context)


def scholarship_list(request):
    scholarships = ScholarshipProgram.objects.all().order_by('-created_at')
    return render(request, 'sdash.html', {
        'scholarships': scholarships,
        'now': timezone.now()
    })
    
    

@login_required
def add_cbt_exam(request):
    """
    View to create a new CBT Exam (title + courses)
    Only accessible by Admin and Moderator.
    """
    # Check role
    if not hasattr(request.user, 'profile') or request.user.profile.role not in ['Admin', 'Moderator']:
        return redirect('home')  # or render a "Permission Denied" page

    error = None
    courses = Course.objects.all()

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        selected = request.POST.getlist('courses')  # get multiple selected courses

        # Validation
        if not title:
            error = "CBT Exam title is required."
        elif CBTExam.objects.filter(title__iexact=title).exists():
            error = "A CBT Exam with this title already exists."
        else:
            # Determine courses to assign
            if 'all' in selected or not selected:
                courses_to_assign = courses
            else:
                courses_to_assign = courses.filter(id__in=selected)

            # Create new CBTExam and assign courses
            new_exam = CBTExam.objects.create(title=title)
            new_exam.courses.set(courses_to_assign)

            return redirect('scholarship_list')  # redirect to your list page

    context = {
        'courses': courses,
        'error': error,
        'now': timezone.now(),
    }
    return render(request, 'add_cbt_exam.html', context)

def edit_scholarship(request, scholarship_id):
    """
    View to edit an existing scholarship
    """
    scholarship = get_object_or_404(ScholarshipProgram, id=scholarship_id)
    courses = Course.objects.all()
    cbt_exams = CBTExam.objects.all()
    error = None

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        passing_score = request.POST.get('passing_score', '').strip()
        exam_date = request.POST.get('exam_date', '').strip()
        cbt_exam_id = request.POST.get('cbt_exam', '').strip()
        course_id = request.POST.get('course', '').strip()
        requirements = request.POST.get('requirements', '').strip()
        tquestion= request.POST.get('tquestion', '').strip()
        exam_duration=request.POST.get('exam_duration').strip()
        exam_end=request.POST.get('exam_end').strip()
        exam_start=request.POST.get('exam_start').strip()
        # Unique title check (exclude current scholarship)
        if ScholarshipProgram.objects.filter(title__iexact=title).exclude(id=scholarship.id).exists():
            error = "A scholarship with this title already exists."
        else:
            scholarship.title = title
            scholarship.passing_score = passing_score
            scholarship.exam_date = exam_date
            scholarship.cbt_exam = CBTExam.objects.get(id=cbt_exam_id)
            scholarship.course = Course.objects.get(id=course_id) if course_id else None
            scholarship.requirements = requirements
            scholarship.tquestion= tquestion
            scholarship.examduration= exam_duration
            scholarship.endtime= exam_end
            scholarship.starttime= exam_start
            scholarship.save()
            return redirect('scholarship_list')

    context = {
        'scholarship': scholarship,
        'courses': courses,
        'cbt_exams': cbt_exams,
        'error': error,
        'now': timezone.now(),
    }
    return render(request, 'edit_scholarship.html', context)


from django.shortcuts import get_object_or_404, redirect
from .models import ScholarshipProgram

def delete_scholarship(request, scholarship_id):
    """
    View to delete a scholarship
    """
    scholarship = get_object_or_404(ScholarshipProgram, id=scholarship_id)

    if request.method == 'POST':
        scholarship.delete()
        return redirect('scholarship_list')  # redirect to scholarship list page
    
def view_cbt_exam(request):
    """
    View to display all CBT Exams with their associated courses
    """
    exams = CBTExam.objects.prefetch_related('courses').all().order_by('title')

    context = {
        'exams': exams
    }
    return render(request, 'view_cbt_exam.html', context)



def edit_cbt_exam(request, exam_id):
    exam = get_object_or_404(CBTExam, id=exam_id)
    error = None
    courses = Course.objects.all()
    selected_courses = [c.id for c in exam.courses.all()]

    if request.method == 'POST':
        title = request.POST.get('title', '').strip()
        selected = request.POST.getlist('courses')
        if not title:
            error = "CBT Exam title is required."
        elif CBTExam.objects.filter(title__iexact=title).exclude(id=exam.id).exists():
            error = "A CBT Exam with this title already exists."
        else:
            exam.title = title
            exam.save()
            if 'all' in selected or not selected:
                courses_to_assign = courses
            else:
                courses_to_assign = courses.filter(id__in=selected)
            exam.courses.set(courses_to_assign)
            return redirect('view_cbt_exam')

    context = {
        'exam': exam,
        'courses': courses,
        'selected_courses': selected_courses,
        'error': error,
        'now': timezone.now(),
    }
    return render(request, 'edit_cbt_exam.html', context)


@login_required
def delete_cbt_exam(request, exam_id):
    """
    Delete a CBT Exam. Only Admin can perform this action.
    """
    # Check role
    if not hasattr(request.user, 'profile') or request.user.profile.role != "Admin":
        messages.error(request, "You do not have permission to delete this CBT Exam.")
        return redirect('view_cbt_exam')

    exam = get_object_or_404(CBTExam, id=exam_id)
    exam.delete()
    messages.success(request, "CBT Exam deleted successfully!")
    return redirect('view_cbt_exam')

# View all CBT Exams
def view_cbt_exam(request):
    exams = CBTExam.objects.prefetch_related('courses').all().order_by('title')
    context = {'exams': exams, 'now': timezone.now()}
    return render(request, 'view_cbt_exam.html', context)


def course_count(request):
    """
    Return the total number of courses.
    """
    total_courses = Course.objects.count()
    return JsonResponse({'total_courses': total_courses})


def question_count(request):
    """
    Return the total number of courses.
    """
    total_question = Question.objects.count()
    return JsonResponse({'total_question': total_question})

def user_logout(request):
    logout(request)
    return redirect('login')  # Redirect to login page after logout


def reset_password(request, user_id):
    """
    Reset a particular user's password to the default value.
    """
    if request.method == "POST":
        user = get_object_or_404(User, id=user_id)

        # Reset password to default
        default_password = "Password@1@"
        user.set_password(default_password)
        user.save()

        messages.success(request, f"Password for {user.first_name} {user.last_name} has been reset to default.")
        return redirect("users_list")  # Adjust to your users list view name
    
    
def change_password(request):
    if request.method == 'POST':
        form = PasswordChangeForm(request.user, request.POST)
        if form.is_valid():
            user = form.save()
            # Keep the user logged in after password change
            update_session_auth_hash(request, user)
            messages.success(request, "✅ Your password has been changed successfully!")
            return redirect('home')  # redirect to dashboard or home
        else:
            messages.error(request, "⚠️ Please correct the errors below.")
    else:
        form = PasswordChangeForm(request.user)
    return render(request, 'change_password.html', {'form': form})



def view_courses(request):
    courses = Course.objects.all()
    return render(request, 'view_courses.html', {'courses': courses})



def edit_course(request, course_id):
    course = get_object_or_404(Course, id=course_id)
    
    if request.method == "POST":
        title = request.POST.get("title")
        if title:
            course.title = title
            course.save()
            messages.success(request, "Course updated successfully!")
            return redirect("view_courses")
        else:
            messages.error(request, "Course title cannot be empty.")
    
    return render(request, "edit_course.html", {"course": course})


@login_required
def delete_course(request, course_id):
    # Only Admin can delete
    if not hasattr(request.user, 'profile') or request.user.profile.role != "Admin":
        messages.error(request, "You do not have permission to delete this course.")
        return redirect("view_courses")
    
    course = get_object_or_404(Course, id=course_id)
    course.delete()
    messages.success(request, "Course deleted successfully!")
    return redirect("view_courses")


def scholarview(request):
    scholarships = ScholarshipProgram.objects.all().order_by('-created_at')
    return render(request, 'appdash.html', {
        'scholarships': scholarships,
        'now': timezone.now()
    })
    

def resultdash(request):
    mesg=' All Wards'
    wardname='0'
    scholarships = ScholarshipProgram.objects.all().order_by('-created_at')
    return render(request, 'resultappdash.html', {
        'scholarships': scholarships,'wardname':wardname,'mesg':mesg,
        'now': timezone.now()
    })
    

def scholarrward(request):
    wardname = request.POST.get('wardname')
    if wardname == '':
        mesg=' All Ward'
    else:
        mesg = "ward " + wardname
    scholarships = ScholarshipProgram.objects.all().order_by('-created_at')
    return render(request, 'resultappdash.html', {
        'scholarships': scholarships,'wardname':wardname,'mesg':mesg,
        'now': timezone.now()
    })
    
# def register_student(request, scholarship_id):
#     scholarship = get_object_or_404(ScholarshipProgram, id=scholarship_id)

#     if request.method == "POST":
#         data = request.POST
#         files = request.FILES

#         ScholarshipApplicant.objects.create(
#             scholarship=scholarship,  # Link applicant to this scholarship
#             # Personal Info 
#             name=data.get("name"),
#             country="Nigeria",
#             gender=data.get("gender"),
#             maritalstatus=data.get("maritalstatus"),
#             phoneno=data.get("phoneno"),
#             email=data.get("email"),
#             address=data.get("address"),
#             city=data.get("city"),
#             state=data.get("state"),
#             lga=data.get("lga"),
#             zipcode=data.get("zipcode"),
#             dob=data.get("dob"),

#             # Academic Info
#             school_name=data.get("school_name"),
#             level_of_study=data.get("level_of_study"),
#             field_of_study=data.get("field_of_study"),
#             gpa=data.get("gpa") if data.get("gpa") else None,
#             graduation_year=data.get("graduation_year") if data.get("graduation_year") else None,
#             scholarship_type=data.get("scholarship_type"),
#             essay=data.get("essay"),
#             previous_awards=data.get("previous_awards"),

#             # Referee Info
#             referee_name=data.get("referee_name"),
#             referee_contact=data.get("referee_contact"),

#             # Documents
#             result=files.get("Result"),
#             passport_photo=files.get("passport_photo"),
#             hobbies=data.get("hobbies"),
#         )

#         messages.success(request, "✅ Application submitted successfully!")
#         return redirect("scholarview")

#     context = {
#         "scholarship": scholarship
#     }
#     return render(request, "app.html", context)

def register_student(request, scholarship_id):
    scholarship = get_object_or_404(ScholarshipProgram, id=scholarship_id)

    if request.method == "POST":
        data = request.POST
        files = request.FILES

        with transaction.atomic():
            # Always fetch the current max id inside the transaction
            last_id = User.objects.aggregate(Max("id"))["id__max"] or 0
            next_id = last_id + 1
            username = f"STU{next_id:03d}"  # STU001, STU002 ...

            # If already taken in User table, keep incrementing until free
            while User.objects.filter(username=username).exists():
                next_id += 10
                username = f"STU{next_id:03d}"

            # Generate random 4-digit password
            password = str(random.randint(1000, 9999))

            # Create User
            user = User.objects.create_user(username=username, password=password)
            user.first_name = data.get("name")
            user.email = data.get("email")
            user.is_staff = False
            user.save()

            # Create Profile
            Profile.objects.create(
                user=user,
                role='Student',
                gender=data.get("gender"),
                marital_status=data.get("maritalstatus"),
                phone=data.get("phoneno"),
                address=data.get("address"),
            )

            # Create Scholarship Applicant (store username + plain password)
            applicant = ScholarshipApplicant.objects.create(
                name=data.get("name"),
                user=user,
                scholarship_type=scholarship_id,
                country="Nigeria",
                gender=data.get("gender"),
                maritalstatus=data.get("maritalstatus"),
                phoneno=data.get("phoneno"),
                email=data.get("email"),
                address=data.get("address"),
                city=data.get("city"),
                state=data.get("state"),
                lga=data.get("lga"),
                zipcode=data.get("zipcode"),
                dob=data.get("dob"),

                school_name=data.get("school_name"),
                level_of_study=data.get("level_of_study"),
                field_of_study=data.get("field_of_study"),
                gpa=data.get("gpa") if data.get("gpa") else None,
                graduation_year=data.get("graduation_year") if data.get("graduation_year") else None,
                essay=data.get("essay"),
                previous_awards=data.get("previous_awards"),
                wardname=data.get("ward"),
                community=data.get("community"),
                referee_name=data.get("referee_name"),
                referee_contact=data.get("referee_contact"),

                result=files.get("result"),
                passport_photo=files.get("passport_photo"),
                idcard=files.get("idcard"),
                lgacert=files.get("lgacert"),
                hobbies=data.get("hobbies"),

                #username=username,
                #plain_password=password,
            )

        messages.success(
            request,
            f"✅ Application submitted! Username: {username}, Password: {password}"
        )
        return redirect("scholarview")

    context = {"scholarship": scholarship}
    return render(request, "app.html", context)


def view_applicants(request, scholarship_id):
    # Get all applicants for this scholarship, ordered by name
    applicants = ScholarshipApplicant.objects.filter(
        scholarship_type=scholarship_id
    ).order_by("name")

    return render(request, "view_applicants.html", {
        "applicant_users": applicants,  # you can rename to "applicants"
        "now": timezone.now()
    })
    
    
def reset_applicant_password(request, user_id):
    user = get_object_or_404(User, username=user_id)
    new_password = '123456'  #str(random.randint(1000, 9999))
    user.set_password(new_password)
    user.save()
   

    messages.success(request, f"Password reset for {user.username}. New password: {new_password}")
    return redirect(request.META.get('HTTP_REFERER', '/'))


def delete_applicant(request, applicant_id):
    applicant = get_object_or_404(ScholarshipApplicant, id=applicant_id)
    applicant.delete()
    messages.success(request, f"Applicant {applicant.name} deleted successfully.")
    return redirect(request.META.get('HTTP_REFERER', '/'))



@login_required
def student_exams_list(request):
    # Get all scholarships
    scholarships = ScholarshipProgram.objects.all()
    
    # Get the IDs of scholarships already taken by this student
    taken_ids = StudentCBTAnswer.objects.filter(student=request.user)\
                                        .values_list('scholarship_id', flat=True)

    return render(request, "student_exams_list.html", {
        "scholarships": scholarships,
        "taken_ids": list(taken_ids),  # convert to list for template lookup
    })
    
    
    
@login_required
def take_exam(request, scholarship_id):
    scholarship = get_object_or_404(ScholarshipProgram, id=scholarship_id)
    
    # Check if student has already taken the exam
    if StudentCBTAnswer.objects.filter(student=request.user, scholarship=scholarship).exists():
        messages.warning(request, "You have already taken this exam. You cannot take it again.")
        return redirect('student_exams_list')
    
    courses = scholarship.cbt_exam.courses.all()
    questions = Question.objects.filter(course__in=courses)
    
    if request.method == "POST":
        for question in questions:
            selected_option = request.POST.get(f"question_{question.id}")
            if selected_option:
                # check if answer is correct
                if selected_option == question.correct_option:
                    score = float(question.mrk)  # use question.mrk as score
                else:
                    score = 0

                StudentCBTAnswer.objects.update_or_create(
                    student=request.user,
                    scholarship=scholarship,
                    question=question,
                    defaults={
                        'selected_option': selected_option,
                        'score': score
                    }
                )
        messages.success(request, "Exam submitted successfully!")
        return redirect('exam_result', scholarship_id=scholarship.id)
    
    return render(request, "cbtexam.html", {"scholarship": scholarship, "questions": questions})
    
    
    
@login_required
def exam_result(request, scholarship_id):
    scholarship = get_object_or_404(ScholarshipProgram, id=scholarship_id)

    # Get student's answers
    answers = StudentCBTAnswer.objects.filter(
        student=request.user,
        scholarship=scholarship
    ).select_related("question")

    # Calculate total and obtained marks
    total_marks = sum(float(ans.question.mrk) for ans in answers)
    obtained_marks = sum(float(ans.score) for ans in answers)
    percentage = (obtained_marks / total_marks * 100) if total_marks else 0

    passing_score = scholarship.passing_score
    status = "PASS" if percentage >= passing_score else "FAILED"

    # Count correct and wrong answers
    correct_count = sum(1 for ans in answers if ans.selected_option == ans.question.correct_option)
    wrong_count = len(answers) - correct_count

    return render(request, "exam_result.html", {
        "scholarship": scholarship,
        "answers": answers,
        "total_score": obtained_marks,
        "total_possible": total_marks,
        "percentage": round(percentage, 2),
        "status": status,
        "passing_score": passing_score,
        "correct_count": correct_count,
        "wrong_count": wrong_count,
    })  
    
    
def edit_applicant(request, applicant_id):
    from django.shortcuts import get_object_or_404, redirect, render
    from django.contrib import messages
    from .models import ScholarshipApplicant
    from .forms import ScholarshipApplicantForm

    applicant = get_object_or_404(ScholarshipApplicant, id=applicant_id)

    if request.method == "POST":
        form = ScholarshipApplicantForm(request.POST, instance=applicant)
        if form.is_valid():
            form.save()
            messages.success(request, "Applicant updated successfully!")
            return redirect("scholarview")  # adjust to your list view
        else:
            messages.error(request, "Please correct the errors below.")
    else:
        form = ScholarshipApplicantForm(instance=applicant)

    wards = range(1, 11)  # Wards 1–10

    return render(request, "edit_applicant.html", {
        "form": form,
        "applicant": applicant,
        "wards": wards,
    })
    
    
    
    
@login_required
def exam_leaderboard(request, scholarship_id, wardname):
    scholarship = get_object_or_404(ScholarshipProgram, id=scholarship_id)

    # Base query: all students who took the exam
    students_answers = StudentCBTAnswer.objects.filter(
        scholarship=scholarship
    ).select_related('student', 'question')

    # Filter by ward if wardname > 0
    if wardname:
        try:
            ward_id = int(wardname)
            if ward_id > 0:
                mesg=" WARD "+wardname
                students_answers = students_answers.filter(
                    student__scholarshipapplicant__wardname=str(ward_id)
                )
            else:
               mesg=' ALL WARDS'
            # ward_id == 0 → no filter, show all
        except ValueError:
            # invalid wardname → no filter, show all
            
            pass

    # Aggregate scores per student
    leaderboard = {}
    for ans in students_answers:
        student_id = ans.student.id
        if student_id not in leaderboard:
            leaderboard[student_id] = {
                'student': ans.student,
                'total_score': 0,
                'total_possible': 0,
            }
        leaderboard[student_id]['total_score'] += float(ans.score)
        leaderboard[student_id]['total_possible'] += float(ans.question.mrk)

    # Calculate percentage and status
    for entry in leaderboard.values():
        total = entry['total_score']
        possible = entry['total_possible']
        entry['percentage'] = round((total / possible * 100) if possible > 0 else 0, 2)
        entry['status'] = "PASS" if entry['percentage'] >= scholarship.passing_score else "FAILED"

    # Sort leaderboard by score descending
    sorted_leaderboard = sorted(
        leaderboard.values(),
        key=lambda x: x['total_score'],
        reverse=True
    )

    return render(request, "exam_leaderboard.html", {
        "scholarship": scholarship,
        "leaderboard": sorted_leaderboard,
        "wardname": wardname,
        "mesg":mesg
    })
    
    


@login_required
def export_leaderboard_excel(request, scholarship_id):
    scholarship = get_object_or_404(ScholarshipProgram, id=scholarship_id)
    students_answers = StudentCBTAnswer.objects.filter(
        scholarship=scholarship
    ).select_related('student', 'question')

    # Aggregate scores per student
    leaderboard = {}
    for ans in students_answers:
        student_id = ans.student.id
        if student_id not in leaderboard:
            leaderboard[student_id] = {
                'student': ans.student,
                'total_score': 0,
                'total_possible': 0
            }
        leaderboard[student_id]['total_score'] += float(ans.score)
        leaderboard[student_id]['total_possible'] += float(ans.question.mrk)

    # Calculate percentage and status
    for entry in leaderboard.values():
        total = entry['total_score']
        possible = entry['total_possible']
        entry['percentage'] = round((total / possible * 100) if possible > 0 else 0, 2)
        entry['status'] = "PASS" if entry['percentage'] >= scholarship.passing_score else "FAILED"

    sorted_leaderboard = sorted(leaderboard.values(), key=lambda x: x['total_score'], reverse=True)

    # Create Excel file in memory
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Leaderboard"

    ws.append(["#", "Student", "Total Score", "Percentage", "Status"])

    for idx, entry in enumerate(sorted_leaderboard, start=1):
        ws.append([
            idx,
            entry['student'].get_full_name(),
            f"{entry['total_score']} / {entry['total_possible']}",
            f"{entry['percentage']}%",
            entry['status']
        ])

    # Save workbook to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    # Return as response
    response = HttpResponse(
        output,
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename=Leaderboard_{scholarship.title}.xlsx'
    return response
    
    
@login_required
def exam_leaderboardpass(request, scholarship_id):
    scholarship = get_object_or_404(ScholarshipProgram, id=scholarship_id)

    # Get all students who took the exam
    students_answers = StudentCBTAnswer.objects.filter(
        scholarship=scholarship
    ).select_related('student', 'question')

    # Aggregate scores per student
    leaderboard = {}
    for ans in students_answers:
        student_id = ans.student.id
        if student_id not in leaderboard:
            leaderboard[student_id] = {
                'student': ans.student,
                'total_score': 0,
                'total_possible': 0
            }
        leaderboard[student_id]['total_score'] += float(ans.score)
        leaderboard[student_id]['total_possible'] += float(ans.question.mrk)

    # Calculate percentage and status
    for entry in leaderboard.values():
        total = entry['total_score']
        possible = entry['total_possible']
        entry['percentage'] = round((total / possible * 100) if possible > 0 else 0, 2)
        entry['status'] = "PASS" if entry['percentage'] >= scholarship.passing_score else "FAILED"

    # Keep only students who passed
    passed_leaderboard = [entry for entry in leaderboard.values() if entry['status'] == "PASS"]

    # Sort leaderboard by total_score descending
    sorted_leaderboard = sorted(passed_leaderboard, key=lambda x: x['total_score'], reverse=True)

    return render(request, "leader.html", {
        "scholarship": scholarship,
        "leaderboard": sorted_leaderboard,
    })
    
    


def reset_password_view(request):
    scholarships = ScholarshipProgram.objects.all()

    if request.method == "POST":
        scholarship_id = request.POST.get("scholarship")
        password = request.POST.get("password")
        confirm_password = request.POST.get("confirm_password")

        if password != confirm_password:
            messages.error(request, "Passwords do not match.")
            return render(request, "reset_all_passwords.html", {"scholarships": scholarships})

        # Get all applicants for the selected scholarship
        applicants = ScholarshipApplicant.objects.filter(
            scholarship_type=scholarship_id,
            user__isnull=False
        )

        if not applicants.exists():
            messages.warning(request, "No applicants found for this scholarship.")
            return redirect("reset_password_view")

        # Collect all users
        users = [applicant.user for applicant in applicants if applicant.user]

        if not users:
            messages.warning(request, "No valid users found to reset password.")
            return redirect("reset_password_view")

        # Hash the password once
        hashed_password = make_password(password)

        # Bulk update all users' passwords in a single query
        User.objects.filter(id__in=[user.id for user in users]).update(password=hashed_password)

        # Optional: log usernames to file (still optional, requires a loop for logging)
        try:
            log_path = os.path.join(os.path.expanduser("~"), "reset_log.txt")
            with open(log_path, "a") as f:
                for user in users:
                    f.write(f"Reset password for {user.username}\n")
        except Exception:
            pass  # Fail silently if file logging isn't allowed

        messages.success(
            request,
            f"Passwords reset successfully for {len(users)} applicants."
        )
        return redirect("reset_password_view")

    # GET request: render the form
    return render(request, "reset_all_passwords.html", {"scholarships": scholarships})
    
    
def ward_form_view(request):
    # Get unique ward names
    wards = ScholarshipApplicant.objects.values_list('wardname', flat=True).distinct().order_by('wardname')
    return render(request, "ward_form.html", {"wards": wards})
    
    
    
def upload_question(request, course_id):
    # your logic here
    return render(request, "upload_question.html", {"course_id": course_id})
    
    
@login_required
def upload_questions(request, course_id):
    course = get_object_or_404(Course, id=course_id)

    if request.method == "POST" and request.FILES.get("question_file"):
        excel_file = request.FILES["question_file"]
        wb = openpyxl.load_workbook(excel_file)
        sheet = wb.active

        for row in sheet.iter_rows(min_row=2, values_only=True):
            # Excel columns: Question, Option A, Option B, Option C, Option D, Score, Correct Option
            text, option_a, option_b, option_c, option_d, mrk, correct_option = row

            Question.objects.create(
                course_id=course_id,
                text=text,
                option_a=option_a,
                option_b=option_b,
                option_c=option_c,
                option_d=option_d,
                mrk=str(mrk),
                correct_option=correct_option.upper()
            )

        messages.success(request, "Questions uploaded successfully!")
        return redirect("view_questions", course_id=course.id)

    return render(request, "upload_question.html", {"course_id": course.id})